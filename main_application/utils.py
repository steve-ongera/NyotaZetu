"""
Utility functions for security analysis and threat detection
"""
from django.utils import timezone
from django.db.models import Count, Q, Avg
from datetime import timedelta
from collections import defaultdict
import re
from user_agents import parse


class SecurityAnalyzer:
    """
    Advanced security analysis utilities
    """
    
    @staticmethod
    def analyze_user_behavior(user, days=7):
        """
        Analyze user behavior patterns to detect anomalies
        Returns a risk score (0-100)
        """
        from .models import UserURLVisit, LoginAttempt, AuditLog
        
        risk_score = 0
        evidence = []
        
        start_date = timezone.now() - timedelta(days=days)
        
        # 1. Check login patterns
        logins = LoginAttempt.objects.filter(
            username=user.username,
            timestamp__gte=start_date
        )
        
        failed_logins = logins.filter(success=False).count()
        if failed_logins > 5:
            risk_score += 20
            evidence.append(f'{failed_logins} failed login attempts')
        
        # 2. Check multiple IP addresses
        ip_addresses = logins.values('ip_address').distinct().count()
        if ip_addresses > 3:
            risk_score += 15
            evidence.append(f'Access from {ip_addresses} different IP addresses')
        
        # 3. Check unusual access times
        night_access = AuditLog.objects.filter(
            user=user,
            timestamp__gte=start_date,
            timestamp__hour__gte=22
        ).count() + AuditLog.objects.filter(
            user=user,
            timestamp__gte=start_date,
            timestamp__hour__lte=5
        ).count()
        
        if night_access > 10:
            risk_score += 10
            evidence.append(f'{night_access} accesses during unusual hours')
        
        # 4. Check rapid page access
        recent_visits = UserURLVisit.objects.filter(
            user=user,
            visited_at__gte=timezone.now() - timedelta(minutes=5)
        ).count()
        
        if recent_visits > 50:
            risk_score += 25
            evidence.append(f'{recent_visits} page visits in 5 minutes (possible scraping)')
        
        # 5. Check access to sensitive URLs
        sensitive_patterns = ['/admin/', '/api/', '/export/', '/download/']
        sensitive_access = UserURLVisit.objects.filter(
            user=user,
            visited_at__gte=start_date
        )
        
        sensitive_count = 0
        for visit in sensitive_access:
            if any(pattern in visit.url_visit.url_path for pattern in sensitive_patterns):
                sensitive_count += 1
        
        if sensitive_count > 20:
            risk_score += 20
            evidence.append(f'{sensitive_count} accesses to sensitive URLs')
        
        # 6. Check for privilege escalation attempts
        escalation_attempts = AuditLog.objects.filter(
            user=user,
            timestamp__gte=start_date,
            description__icontains='unauthorized'
        ).count()
        
        if escalation_attempts > 0:
            risk_score += 30
            evidence.append(f'{escalation_attempts} unauthorized access attempts')
        
        return {
            'risk_score': min(risk_score, 100),
            'evidence': evidence,
            'confidence': 85.0 if len(evidence) > 2 else 60.0
        }
    
    @staticmethod
    def detect_phishing_indicators(user):
        """
        Detect potential phishing or account compromise indicators
        """
        from .models import UserSession, LoginAttempt
        
        indicators = []
        
        # Check for multiple simultaneous sessions
        active_sessions = UserSession.objects.filter(
            user=user,
            is_active=True
        )
        
        if active_sessions.count() > 2:
            indicators.append({
                'type': 'multiple_devices',
                'severity': 'medium',
                'description': f'{active_sessions.count()} simultaneous active sessions'
            })
        
        # Check for geographically impossible logins
        recent_logins = LoginAttempt.objects.filter(
            username=user.username,
            success=True,
            timestamp__gte=timezone.now() - timedelta(hours=1)
        ).values_list('ip_address', flat=True)
        
        unique_ips = set(recent_logins)
        if len(unique_ips) > 2:
            indicators.append({
                'type': 'impossible_travel',
                'severity': 'high',
                'description': f'Logins from {len(unique_ips)} different locations in 1 hour'
            })
        
        # Check for sudden change in user agent
        recent_sessions = UserSession.objects.filter(
            user=user,
            login_time__gte=timezone.now() - timedelta(days=1)
        ).values_list('user_agent', flat=True)
        
        unique_agents = set(recent_sessions)
        if len(unique_agents) > 3:
            indicators.append({
                'type': 'device_switching',
                'severity': 'medium',
                'description': f'{len(unique_agents)} different devices in 24 hours'
            })
        
        return indicators
    
    @staticmethod
    def analyze_threat_patterns(days=30):
        """
        Analyze security threat patterns over time
        """
        from .models import SecurityThreat
        
        start_date = timezone.now() - timedelta(days=days)
        
        threats = SecurityThreat.objects.filter(
            detected_at__gte=start_date
        )
        
        # Group by threat type
        threat_distribution = threats.values('threat_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Group by IP address (identify repeat offenders)
        repeat_offenders = threats.values('ip_address').annotate(
            count=Count('id')
        ).filter(count__gte=3).order_by('-count')
        
        # Severity distribution
        severity_distribution = threats.values('severity').annotate(
            count=Count('id')
        )
        
        # Trend analysis (daily counts)
        daily_threats = threats.extra(
            select={'day': 'date(detected_at)'}
        ).values('day').annotate(
            count=Count('id')
        ).order_by('day')
        
        return {
            'total_threats': threats.count(),
            'unresolved': threats.filter(resolved=False).count(),
            'threat_distribution': list(threat_distribution),
            'repeat_offenders': list(repeat_offenders),
            'severity_distribution': list(severity_distribution),
            'daily_trend': list(daily_threats),
        }
    
    @staticmethod
    def parse_user_agent(user_agent_string):
        """
        Parse user agent string to extract device info
        """
        try:
            user_agent = parse(user_agent_string)
            
            return {
                'device_type': 'Mobile' if user_agent.is_mobile else 'Tablet' if user_agent.is_tablet else 'Desktop',
                'browser': f"{user_agent.browser.family} {user_agent.browser.version_string}",
                'os': f"{user_agent.os.family} {user_agent.os.version_string}",
                'is_bot': user_agent.is_bot,
            }
        except:
            return {
                'device_type': 'Unknown',
                'browser': 'Unknown',
                'os': 'Unknown',
                'is_bot': False,
            }
    
    @staticmethod
    def calculate_session_risk(session):
        """
        Calculate risk score for a user session
        """
        from .models import LoginAttempt, SecurityThreat
        
        risk_score = 0
        
        # Check for recent threats from same IP
        recent_threats = SecurityThreat.objects.filter(
            ip_address=session.ip_address,
            detected_at__gte=timezone.now() - timedelta(days=7)
        ).count()
        
        if recent_threats > 0:
            risk_score += 30
        
        # Check for failed login attempts
        failed_attempts = LoginAttempt.objects.filter(
            ip_address=session.ip_address,
            success=False,
            timestamp__gte=timezone.now() - timedelta(hours=24)
        ).count()
        
        if failed_attempts > 3:
            risk_score += 20
        
        # Check session duration
        session_duration = timezone.now() - session.login_time
        if session_duration > timedelta(hours=8):
            risk_score += 10
        
        # Check for unusual user agent
        device_info = SecurityAnalyzer.parse_user_agent(session.user_agent)
        if device_info['is_bot']:
            risk_score += 40
        
        return min(risk_score, 100)
    
    @staticmethod
    def generate_security_report(days=30):
        """
        Generate comprehensive security report
        """
        from .models import (
            SecurityThreat, SuspiciousActivity, LoginAttempt,
            User, UserSession, AuditLog
        )
        
        start_date = timezone.now() - timedelta(days=days)
        
        report = {
            'period': {
                'start': start_date,
                'end': timezone.now(),
                'days': days,
            },
            'threats': {
                'total': SecurityThreat.objects.filter(detected_at__gte=start_date).count(),
                'critical': SecurityThreat.objects.filter(
                    detected_at__gte=start_date,
                    severity='critical'
                ).count(),
                'resolved': SecurityThreat.objects.filter(
                    detected_at__gte=start_date,
                    resolved=True
                ).count(),
                'unresolved': SecurityThreat.objects.filter(
                    detected_at__gte=start_date,
                    resolved=False
                ).count(),
            },
            'suspicious_activities': {
                'total': SuspiciousActivity.objects.filter(detected_at__gte=start_date).count(),
                'high_risk': SuspiciousActivity.objects.filter(
                    detected_at__gte=start_date,
                    risk_score__gte=70
                ).count(),
                'investigated': SuspiciousActivity.objects.filter(
                    detected_at__gte=start_date,
                    investigated=True
                ).count(),
            },
            'authentication': {
                'successful_logins': LoginAttempt.objects.filter(
                    timestamp__gte=start_date,
                    success=True
                ).count(),
                'failed_attempts': LoginAttempt.objects.filter(
                    timestamp__gte=start_date,
                    success=False
                ).count(),
                'unique_users': LoginAttempt.objects.filter(
                    timestamp__gte=start_date,
                    success=True
                ).values('username').distinct().count(),
            },
            'user_activity': {
                'total_users': User.objects.count(),
                'active_users': User.objects.filter(last_login__gte=start_date).count(),
                'new_users': User.objects.filter(date_joined__gte=start_date).count(),
                'current_sessions': UserSession.objects.filter(is_active=True).count(),
            },
            'audit': {
                'total_actions': AuditLog.objects.filter(timestamp__gte=start_date).count(),
                'security_events': AuditLog.objects.filter(
                    timestamp__gte=start_date,
                    action='security_event'
                ).count(),
            }
        }
        
        # Calculate metrics
        total_logins = report['authentication']['successful_logins'] + report['authentication']['failed_attempts']
        if total_logins > 0:
            report['authentication']['success_rate'] = round(
                (report['authentication']['successful_logins'] / total_logins) * 100, 2
            )
        else:
            report['authentication']['success_rate'] = 100.0
        
        # Threat resolution rate
        total_threats = report['threats']['total']
        if total_threats > 0:
            report['threats']['resolution_rate'] = round(
                (report['threats']['resolved'] / total_threats) * 100, 2
            )
        else:
            report['threats']['resolution_rate'] = 100.0
        
        return report


class ThreatDetector:
    """
    Real-time threat detection utilities
    """
    
    @staticmethod
    def detect_sql_injection(data):
        """Detect SQL injection patterns"""
        sql_patterns = [
            r"union.*select",
            r"insert.*into",
            r"delete.*from",
            r"drop.*table",
            r"update.*set",
            r"'.*or.*'.*=.*'",
            r"exec\s*\(",
        ]
        
        for pattern in sql_patterns:
            if re.search(pattern, str(data), re.IGNORECASE):
                return True
        return False
    
    @staticmethod
    def detect_xss(data):
        """Detect XSS patterns"""
        xss_patterns = [
            r"<script.*?>",
            r"javascript:",
            r"onerror\s*=",
            r"onload\s*=",
            r"onclick\s*=",
            r"<iframe.*?>",
        ]
        
        for pattern in xss_patterns:
            if re.search(pattern, str(data), re.IGNORECASE):
                return True
        return False
    
    @staticmethod
    def detect_path_traversal(data):
        """Detect path traversal attempts"""
        if '../' in str(data) or '..\\' in str(data):
            return True
        return False
    
    @staticmethod
    def detect_command_injection(data):
        """Detect command injection patterns"""
        cmd_patterns = [
            r";\s*(ls|cat|echo|wget|curl)",
            r"\|\s*(ls|cat|echo|wget|curl)",
            r"&&\s*(ls|cat|echo|wget|curl)",
            r"`.*`",
            r"\$\(.*\)",
        ]
        
        for pattern in cmd_patterns:
            if re.search(pattern, str(data), re.IGNORECASE):
                return True
        return False
    
    @staticmethod
    def analyze_request(request):
        """
        Analyze request for all threat types
        """
        threats = []
        
        # Check URL
        full_path = request.get_full_path()
        
        if ThreatDetector.detect_sql_injection(full_path):
            threats.append('sql_injection')
        if ThreatDetector.detect_xss(full_path):
            threats.append('xss')
        if ThreatDetector.detect_path_traversal(full_path):
            threats.append('path_traversal')
        if ThreatDetector.detect_command_injection(full_path):
            threats.append('code_injection')
        
        # Check POST data
        if request.method == 'POST':
            try:
                body = request.body.decode('utf-8')
                
                if ThreatDetector.detect_sql_injection(body):
                    threats.append('sql_injection')
                if ThreatDetector.detect_xss(body):
                    threats.append('xss')
                if ThreatDetector.detect_command_injection(body):
                    threats.append('code_injection')
            except:
                pass
        
        return list(set(threats))  # Remove duplicates
    
    
    
"""
Decorators and Utility Functions for Constituency Admin
"""

from django.shortcuts import redirect
from django.contrib import messages
from functools import wraps
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
import requests


def constituency_admin_required(view_func):
    """
    Decorator to ensure user is a constituency admin
    """
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, "Please log in to access this page.")
            return redirect('login')
        
        if request.user.user_type != 'constituency_admin':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('home')
        
        return view_func(request, *args, **kwargs)
    
    return wrapper


def create_audit_log(user, action, table_affected, record_id, description, request):
    """
    Create an audit log entry
    """
    from .models import AuditLog
    
    # Get IP address
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip_address = x_forwarded_for.split(',')[0]
    else:
        ip_address = request.META.get('REMOTE_ADDR')
    
    # Get user agent
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    
    AuditLog.objects.create(
        user=user,
        action=action,
        table_affected=table_affected,
        record_id=str(record_id) if record_id else None,
        description=description,
        ip_address=ip_address,
        user_agent=user_agent
    )


def send_sms_notification(phone_number, message):
    """
    Send SMS notification via SMS gateway
    Configure your SMS provider here (Africa's Talking, Twilio, etc.)
    """
    from .models import SMSLog
    
    try:
        # Example using Africa's Talking (configure your credentials)
        # You'll need to install africas_talking package
        # pip install africastalking
        
        # import africastalking
        # africastalking.initialize(
        #     username=settings.AFRICASTALKING_USERNAME,
        #     api_key=settings.AFRICASTALKING_API_KEY
        # )
        # sms = africastalking.SMS
        # response = sms.send(message, [phone_number])
        
        # For now, just log the SMS
        sms_log = SMSLog.objects.create(
            phone_number=phone_number,
            message=message,
            status='sent'  # Change to 'pending' in production
        )
        
        return True
    
    except Exception as e:
        # Log failed SMS
        SMSLog.objects.create(
            phone_number=phone_number,
            message=message,
            status='failed',
            gateway_response=str(e)
        )
        return False


def send_email_notification(email_address, subject, message):
    """
    Send email notification
    """
    from .models import EmailLog
    
    try:
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email_address],
            fail_silently=False,
        )
        
        EmailLog.objects.create(
            email_address=email_address,
            subject=subject,
            message=message,
            status='sent'
        )
        
        return True
    
    except Exception as e:
        EmailLog.objects.create(
            email_address=email_address,
            subject=subject,
            message=message,
            status='failed'
        )
        return False


def calculate_priority_score(application):
    """
    Calculate priority score for application ranking
    Based on multiple factors
    """
    score = 0
    
    # Orphan status (high priority)
    if application.is_total_orphan:
        score += 30
    elif application.is_orphan:
        score += 20
    
    # Special needs
    if application.is_disabled:
        score += 15
    
    # Chronic illness
    if application.has_chronic_illness:
        score += 10
    
    # Household income (lower income = higher score)
    if application.household_monthly_income:
        if application.household_monthly_income < 10000:
            score += 20
        elif application.household_monthly_income < 20000:
            score += 15
        elif application.household_monthly_income < 30000:
            score += 10
    
    # Number of siblings in school
    if application.number_of_siblings_in_school >= 4:
        score += 15
    elif application.number_of_siblings_in_school >= 2:
        score += 10
    
    # Academic performance (merit component)
    if application.previous_academic_year_average:
        if application.previous_academic_year_average >= 70:
            score += 10
        elif application.previous_academic_year_average >= 60:
            score += 5
    
    # Fee balance (higher balance = higher need)
    if application.fees_balance >= 50000:
        score += 10
    elif application.fees_balance >= 30000:
        score += 5
    
    return min(score, 100)  # Cap at 100


def get_constituency_budget_status(constituency, fiscal_year):
    """
    Get budget utilization status for constituency
    """
    from .models import Application, Allocation
    from django.db.models import Sum
    
    cdf_budget = constituency.cdf_bursary_allocation or 0
    
    applications = Application.objects.filter(
        applicant__constituency=constituency,
        bursary_source__in=['cdf', 'both'],
        fiscal_year=fiscal_year
    )
    
    total_allocated = Allocation.objects.filter(
        application__in=applications
    ).aggregate(Sum('amount_allocated'))['amount_allocated__sum'] or 0
    
    utilization_rate = (total_allocated / cdf_budget * 100) if cdf_budget > 0 else 0
    
    return {
        'budget': cdf_budget,
        'allocated': total_allocated,
        'available': cdf_budget - total_allocated,
        'utilization_rate': utilization_rate
    }


def export_applications_to_excel(applications, filename='applications.xlsx'):
    """
    Export applications to Excel file
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Applications'
    
    # Headers
    headers = [
        'Application No.', 'Name', 'ID Number', 'Gender', 'Ward',
        'Institution', 'Course', 'Year', 'Amount Requested',
        'Status', 'Date Submitted'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row, app in enumerate(applications, 2):
        ws.cell(row=row, column=1, value=app.application_number)
        ws.cell(row=row, column=2, value=app.applicant.user.get_full_name())
        ws.cell(row=row, column=3, value=app.applicant.id_number)
        ws.cell(row=row, column=4, value=app.applicant.get_gender_display())
        ws.cell(row=row, column=5, value=app.applicant.ward.name if app.applicant.ward else 'N/A')
        ws.cell(row=row, column=6, value=app.institution.name)
        ws.cell(row=row, column=7, value=app.course_name or 'N/A')
        ws.cell(row=row, column=8, value=app.year_of_study)
        ws.cell(row=row, column=9, value=float(app.amount_requested))
        ws.cell(row=row, column=10, value=app.get_status_display())
        ws.cell(row=row, column=11, value=app.date_submitted.strftime('%Y-%m-%d') if app.date_submitted else 'N/A')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    
    return response


def generate_payment_voucher(allocation):
    """
    Generate payment voucher for allocation
    Returns PDF file
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib.styles import getSampleStyleSheet
    from io import BytesIO
    from django.http import HttpResponse
    
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width/2, height - 1*inch, "CONSTITUENCY DEVELOPMENT FUND")
    p.drawCentredString(width/2, height - 1.3*inch, f"{allocation.application.applicant.constituency.name.upper()} CONSTITUENCY")
    
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width/2, height - 1.7*inch, "BURSARY PAYMENT VOUCHER")
    
    # Voucher details
    y_position = height - 2.3*inch
    p.setFont("Helvetica", 11)
    
    details = [
        ("Voucher No:", allocation.cheque_number or 'N/A'),
        ("Date:", allocation.allocation_date.strftime('%d/%m/%Y')),
        ("Payee:", allocation.application.institution.name),
        ("Student Name:", allocation.application.applicant.user.get_full_name()),
        ("ID Number:", allocation.application.applicant.id_number),
        ("Admission No:", allocation.application.admission_number),
        ("Course:", allocation.application.course_name or 'N/A'),
        ("Year of Study:", str(allocation.application.year_of_study)),
        ("Amount:", f"KES {allocation.amount_allocated:,.2f}"),
    ]
    
    for label, value in details:
        p.setFont("Helvetica-Bold", 11)
        p.drawString(1*inch, y_position, label)
        p.setFont("Helvetica", 11)
        p.drawString(2.5*inch, y_position, str(value))
        y_position -= 0.3*inch
    
    # Signatures
    y_position -= 0.5*inch
    p.line(1*inch, y_position, 3.5*inch, y_position)
    p.drawString(1*inch, y_position - 0.2*inch, "Prepared by")
    
    p.line(4.5*inch, y_position, 7*inch, y_position)
    p.drawString(4.5*inch, y_position - 0.2*inch, "Approved by")
    
    # Footer
    p.setFont("Helvetica", 9)
    p.drawCentredString(width/2, 0.5*inch, "This is a computer-generated document")
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    return buffer


def validate_application_documents(application):
    """
    Validate if application has all required documents
    """
    from .models import Document
    
    required_docs = [
        'id_card',
        'admission_letter',
        'fee_structure',
    ]
    
    uploaded_docs = application.documents.filter(
        document_type__in=required_docs
    ).values_list('document_type', flat=True)
    
    missing_docs = set(required_docs) - set(uploaded_docs)
    
    return {
        'is_complete': len(missing_docs) == 0,
        'missing_documents': list(missing_docs),
        'total_uploaded': application.documents.count()
    }


def get_ward_allocation_summary(ward, fiscal_year):
    """
    Get allocation summary for a specific ward
    """
    from .models import Application, Allocation
    from django.db.models import Sum, Count
    
    applications = Application.objects.filter(
        applicant__ward=ward,
        fiscal_year=fiscal_year
    )
    
    allocations = Allocation.objects.filter(
        application__in=applications
    )
    
    summary = {
        'total_applications': applications.count(),
        'approved_count': applications.filter(status='approved').count(),
        'total_requested': applications.aggregate(Sum('amount_requested'))['amount_requested__sum'] or 0,
        'total_allocated': allocations.aggregate(Sum('amount_allocated'))['amount_allocated__sum'] or 0,
        'male_beneficiaries': applications.filter(
            applicant__gender='M',
            status__in=['approved', 'disbursed']
        ).count(),
        'female_beneficiaries': applications.filter(
            applicant__gender='F',
            status__in=['approved', 'disbursed']
        ).count(),
    }
    
    return summary


def check_duplicate_application(applicant, fiscal_year):
    """
    Check if applicant has already applied in the same fiscal year
    """
    from .models import Application
    
    existing = Application.objects.filter(
        applicant=applicant,
        fiscal_year=fiscal_year,
        bursary_source__in=['cdf', 'both']
    ).exists()
    
    return existing


def get_user_constituency(user):
    """
    Get the constituency assigned to the user based on their role
    Returns: Constituency object or None
    """
    try:
        # For applicants - get from their profile
        if hasattr(user, 'applicant_profile'):
            return user.applicant_profile.constituency
        
        # For other user types, you might need additional logic
        # based on your user model structure
        return None
        
    except Exception as e:
        return None
    
def get_user_ward(user):
    """
    Get the ward assigned to the user based on their role
    Returns: Ward object or None
    """
    try:
        # For applicants - get from their profile
        if hasattr(user, 'applicant_profile'):
            return user.applicant_profile.ward
        
        # For other user types
        return None
        
    except Exception as e:
        return None
    
    
"""
Utility functions for the bursary management system
"""

from django.utils import timezone
from .models import Notification, AuditLog, EmailLog, SMSLog
from django.core.mail import send_mail
from django.conf import settings


def create_audit_log(user, action, table_affected, description, ip_address, record_id=None, old_values=None, new_values=None):
    """
    Create an audit log entry
    
    Args:
        user: User performing the action
        action: Type of action (create, update, delete, view, etc.)
        table_affected: Name of the database table/model
        description: Human-readable description
        ip_address: IP address of the user
        record_id: ID of the affected record (optional)
        old_values: Dict of old values for updates (optional)
        new_values: Dict of new values for updates (optional)
    """
    try:
        AuditLog.objects.create(
            user=user,
            action=action,
            table_affected=table_affected,
            record_id=str(record_id) if record_id else None,
            description=description,
            ip_address=ip_address,
            old_values=old_values,
            new_values=new_values
        )
    except Exception as e:
        print(f"Error creating audit log: {e}")


def send_notification(user, notification_type, title, message, related_application=None):
    """
    Send in-app notification to user
    
    Args:
        user: User to notify
        notification_type: Type of notification
        title: Notification title
        message: Notification message
        related_application: Related application object (optional)
    """
    try:
        Notification.objects.create(
            user=user,
            notification_type=notification_type,
            title=title,
            message=message,
            related_application=related_application
        )
    except Exception as e:
        print(f"Error creating notification: {e}")


def send_email_notification(user, subject, message, related_application=None):
    """
    Send email notification to user
    
    Args:
        user: User to email
        subject: Email subject
        message: Email message
        related_application: Related application object (optional)
    """
    try:
        # Send actual email
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=False,
        )
        
        # Log email
        EmailLog.objects.create(
            recipient=user,
            email_address=user.email,
            subject=subject,
            message=message,
            related_application=related_application,
            status='sent'
        )
    except Exception as e:
        print(f"Error sending email: {e}")
        
        # Log failed email
        EmailLog.objects.create(
            recipient=user,
            email_address=user.email,
            subject=subject,
            message=message,
            related_application=related_application,
            status='failed'
        )


def send_sms_notification(user, message, related_application=None):
    """
    Send SMS notification to user
    
    Args:
        user: User to SMS
        message: SMS message
        related_application: Related application object (optional)
    """
    try:
        phone_number = user.phone_number
        
        if not phone_number:
            print(f"User {user.username} has no phone number")
            return
        
        # TODO: Integrate with actual SMS gateway (Africa's Talking, etc.)
        # For now, just log it
        
        SMSLog.objects.create(
            recipient=user,
            phone_number=phone_number,
            message=message,
            related_application=related_application,
            status='sent'  # Change to 'pending' when actual integration is done
        )
    except Exception as e:
        print(f"Error sending SMS: {e}")
        
        # Log failed SMS
        SMSLog.objects.create(
            recipient=user,
            phone_number=user.phone_number if hasattr(user, 'phone_number') else '',
            message=message,
            related_application=related_application,
            status='failed'
        )


def calculate_priority_score(application):
    """
    Calculate priority score for an application based on various factors
    
    Args:
        application: Application object
        
    Returns:
        Decimal: Priority score (0-100)
    """
    from decimal import Decimal
    
    score = Decimal('0.0')
    
    # Vulnerability factors (40 points max)
    if application.is_total_orphan:
        score += Decimal('15.0')
    elif application.is_orphan:
        score += Decimal('10.0')
    
    if application.is_disabled:
        score += Decimal('10.0')
    
    if application.has_chronic_illness:
        score += Decimal('5.0')
    
    if application.applicant.special_needs:
        score += Decimal('10.0')
    
    # Financial need (30 points max)
    if application.household_monthly_income:
        if application.household_monthly_income < 5000:
            score += Decimal('15.0')
        elif application.household_monthly_income < 10000:
            score += Decimal('10.0')
        elif application.household_monthly_income < 20000:
            score += Decimal('5.0')
    
    # Family size consideration
    if application.number_of_siblings > 5:
        score += Decimal('10.0')
    elif application.number_of_siblings > 3:
        score += Decimal('5.0')
    
    # Fees balance (15 points max)
    if application.fees_balance:
        total_fees = application.total_fees_payable or Decimal('1.0')
        balance_ratio = application.fees_balance / total_fees
        
        if balance_ratio > 0.8:  # More than 80% unpaid
            score += Decimal('15.0')
        elif balance_ratio > 0.5:  # More than 50% unpaid
            score += Decimal('10.0')
        elif balance_ratio > 0.3:  # More than 30% unpaid
            score += Decimal('5.0')
    
    # Academic performance (15 points max) - bonus for merit
    if application.previous_academic_year_average:
        if application.previous_academic_year_average >= 80:
            score += Decimal('15.0')
        elif application.previous_academic_year_average >= 70:
            score += Decimal('10.0')
        elif application.previous_academic_year_average >= 60:
            score += Decimal('5.0')
    
    # Cap at 100
    if score > 100:
        score = Decimal('100.0')
    
    return score


def get_user_ip(request):
    """
    Get user's IP address from request
    
    Args:
        request: Django request object
        
    Returns:
        str: IP address
    """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def format_currency(amount):
    """
    Format decimal amount as Kenyan Shillings
    
    Args:
        amount: Decimal or float amount
        
    Returns:
        str: Formatted currency string
    """
    if amount is None:
        return "KES 0.00"
    
    return f"KES {amount:,.2f}"


def get_fiscal_year_display(fiscal_year):
    """
    Get display string for fiscal year
    
    Args:
        fiscal_year: FiscalYear object
        
    Returns:
        str: Display string
    """
    if fiscal_year:
        return f"{fiscal_year.name} ({fiscal_year.start_date.strftime('%b %Y')} - {fiscal_year.end_date.strftime('%b %Y')})"
    return "No Fiscal Year"


def check_application_eligibility(applicant, fiscal_year):
    """
    Check if applicant is eligible to apply for the given fiscal year
    
    Args:
        applicant: Applicant object
        fiscal_year: FiscalYear object
        
    Returns:
        tuple: (is_eligible: bool, reason: str)
    """
    # Check if fiscal year is active
    if not fiscal_year.is_active:
        return False, "The fiscal year is not active"
    
    # Check if applications are open
    if not fiscal_year.application_open:
        return False, "Applications are currently closed"
    
    # Check deadline
    if fiscal_year.application_deadline and timezone.now().date() > fiscal_year.application_deadline:
        return False, "Application deadline has passed"
    
    # Check if applicant is verified
    if not applicant.is_verified:
        return False, "Your profile is not yet verified"
    
    # Check if applicant already has an application for this fiscal year
    from .models import Application
    existing_application = Application.objects.filter(
        applicant=applicant,
        fiscal_year=fiscal_year
    ).exists()
    
    if existing_application:
        return False, "You already have an application for this fiscal year"
    
    return True, "Eligible"


def generate_application_number(county_code, fiscal_year_name):
    """
    Generate unique application number
    
    Args:
        county_code: County code (e.g., "047")
        fiscal_year_name: Fiscal year name (e.g., "2024-2025")
        
    Returns:
        str: Unique application number
    """
    import uuid
    
    year = fiscal_year_name.split('-')[0]
    random_string = uuid.uuid4().hex[:6].upper()
    
    return f"KB-{county_code}-{year}-{random_string}"


def get_review_status_badge(status):
    """
    Get Bootstrap badge class for review status
    
    Args:
        status: Status string
        
    Returns:
        str: Bootstrap badge class
    """
    status_badges = {
        'draft': 'secondary',
        'submitted': 'info',
        'under_review': 'warning',
        'approved': 'success',
        'rejected': 'danger',
        'disbursed': 'primary',
        'pending_documents': 'warning',
    }
    
    return status_badges.get(status, 'secondary')


def get_recommendation_badge(recommendation):
    """
    Get Bootstrap badge class for review recommendation
    
    Args:
        recommendation: Recommendation string
        
    Returns:
        str: Bootstrap badge class
    """
    recommendation_badges = {
        'approve': 'success',
        'reject': 'danger',
        'more_info': 'warning',
        'forward': 'info',
    }
    
    return recommendation_badges.get(recommendation, 'secondary')


def paginate_queryset(queryset, page_number, per_page=20):
    """
    Paginate a queryset
    
    Args:
        queryset: Django queryset
        page_number: Current page number
        per_page: Items per page
        
    Returns:
        Page object
    """
    from django.core.paginator import Paginator
    
    paginator = Paginator(queryset, per_page)
    return paginator.get_page(page_number)